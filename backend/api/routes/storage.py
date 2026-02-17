"""
Storage API — manage workflow files (results, icons, docs, etc.).

Endpoints:
  GET  /storage/workflows/{workflow_id}/files   — List all files
  GET  /storage/files/{path:path}               — Download/serve a file
  POST /storage/workflows/{workflow_id}/init     — Initialize folder structure
  GET  /storage/stats                            — Overall storage statistics
"""

import mimetypes
from pathlib import Path

from fastapi import APIRouter, HTTPException, status, Depends, UploadFile, File, Form
from fastapi.responses import FileResponse
from typing import Optional

from app.dependencies import get_db, get_current_active_user
from core.security import TokenPayload
from services.storage_service import get_storage_service
from services.workflow_service import WorkflowService
from sqlalchemy.ext.asyncio import AsyncSession

import logging

logger = logging.getLogger(__name__)
router = APIRouter(tags=["storage"])


@router.get("/workflows/{workflow_id}/files")
async def list_workflow_files(
    workflow_id: str,
    current_user: TokenPayload = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """List all files in a workflow's storage folder."""
    # Verify workflow belongs to user's org
    svc = WorkflowService(db)
    wf = await svc.get_by_id_and_org(workflow_id, current_user.org_id)
    if not wf:
        raise HTTPException(status_code=404, detail="Workflow not found")

    storage = get_storage_service()
    files = storage.list_workflow_files(workflow_id)

    # If no folder exists yet, init it
    if not files:
        storage.init_workflow_folder(workflow_id, wf.name, wf.description or "")
        files = storage.list_workflow_files(workflow_id)

    return {
        "workflow_id": workflow_id,
        "workflow_name": wf.name,
        "folder": storage.find_workflow_dir(workflow_id).name if storage.find_workflow_dir(workflow_id) else None,
        "files": files,
    }


@router.get("/files/{file_path:path}")
async def serve_file(
    file_path: str,
    current_user: TokenPayload = Depends(get_current_active_user),
):
    """Serve/download a file from storage by its relative path."""
    storage = get_storage_service()
    resolved = storage.get_file_path(file_path)
    if not resolved:
        raise HTTPException(status_code=404, detail="File not found")

    content_type = mimetypes.guess_type(str(resolved))[0] or "application/octet-stream"
    return FileResponse(
        path=str(resolved),
        media_type=content_type,
        filename=resolved.name,
    )


@router.post("/workflows/{workflow_id}/upload/{subfolder}")
async def upload_file(
    workflow_id: str,
    subfolder: str,
    file: UploadFile = File(...),
    current_user: TokenPayload = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload a file to a workflow's subfolder (icons, docs, etc.)."""
    allowed_subfolders = {"icons", "docs", "config", "screenshots"}
    if subfolder not in allowed_subfolders:
        raise HTTPException(status_code=400, detail=f"Invalid subfolder. Allowed: {', '.join(allowed_subfolders)}")

    svc = WorkflowService(db)
    wf = await svc.get_by_id_and_org(workflow_id, current_user.org_id)
    if not wf:
        raise HTTPException(status_code=404, detail="Workflow not found")

    storage = get_storage_service()
    wf_dir = storage.find_workflow_dir(workflow_id)
    if not wf_dir:
        wf_dir = storage.init_workflow_folder(workflow_id, wf.name, wf.description or "")

    target_dir = wf_dir / subfolder
    target_dir.mkdir(exist_ok=True)

    # Sanitize filename
    safe_name = file.filename.replace("/", "_").replace("\\", "_")
    target_path = target_dir / safe_name

    # Write file
    content = await file.read()
    target_path.write_bytes(content)

    logger.info(f"Uploaded {safe_name} to {subfolder}/ for workflow {workflow_id}")
    return {
        "filename": safe_name,
        "subfolder": subfolder,
        "size": len(content),
        "path": str(target_path.relative_to(storage.base_path)),
    }


@router.post("/workflows/{workflow_id}/init")
async def init_workflow_folder(
    workflow_id: str,
    current_user: TokenPayload = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Explicitly initialize the folder structure for a workflow."""
    svc = WorkflowService(db)
    wf = await svc.get_by_id_and_org(workflow_id, current_user.org_id)
    if not wf:
        raise HTTPException(status_code=404, detail="Workflow not found")

    storage = get_storage_service()
    wf_dir = storage.init_workflow_folder(workflow_id, wf.name, wf.description or "")

    return {
        "workflow_id": workflow_id,
        "folder": wf_dir.name,
        "subfolders": [s.name for s in wf_dir.iterdir() if s.is_dir()],
        "message": f"Folder structure initialized at {wf_dir.name}/",
    }


@router.get("/workflows/{workflow_id}/latest-results")
async def get_latest_results(
    workflow_id: str,
    current_user: TokenPayload = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Get the latest execution results for a workflow from DB."""
    import json as _json
    from sqlalchemy import text as sa_text

    svc = WorkflowService(db)
    wf = await svc.get_by_id_and_org(workflow_id, current_user.org_id)
    if not wf:
        raise HTTPException(status_code=404, detail="Workflow not found")

    # Fetch from DB execution_states (authoritative source)
    # Use latest COMPLETED execution — running executions have no state_data yet
    row = (await db.execute(sa_text(
        "SELECT e.id, e.status, e.started_at, e.completed_at, e.duration_ms, "
        "       es.state_data "
        "FROM executions e "
        "LEFT JOIN execution_states es ON es.execution_id = e.id "
        "WHERE e.workflow_id = :wid AND e.status = 'completed' "
        "ORDER BY e.completed_at DESC LIMIT 1"
    ), {"wid": workflow_id})).fetchone()

    if not row or not row[5]:
        raise HTTPException(status_code=404, detail="No results available yet. Run the workflow first.")

    state_data = row[5] if isinstance(row[5], dict) else _json.loads(row[5])

    return {
        "execution_id": str(row[0]),
        "status": row[1],
        "started_at": row[2].isoformat() if row[2] else None,
        "completed_at": row[3].isoformat() if row[3] else None,
        "data": state_data,
    }


@router.get("/workflows/{workflow_id}/latest-results/download")
async def download_latest_results(
    workflow_id: str,
    current_user: TokenPayload = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Download the latest execution results as an Excel (.xlsx) file.
    Fetches full state_data from execution_states DB table,
    extracts actual product/output data from step outputs.
    """
    import json as _json
    import tempfile
    from datetime import datetime
    from sqlalchemy import text as sa_text
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    svc = WorkflowService(db)
    wf = await svc.get_by_id_and_org(workflow_id, current_user.org_id)
    if not wf:
        raise HTTPException(status_code=404, detail="Workflow not found")

    # Get latest COMPLETED execution + its full state_data from DB
    # Running executions have no state_data yet, so we skip them
    row = (await db.execute(sa_text(
        "SELECT e.id, e.status, e.started_at, e.completed_at, e.duration_ms, "
        "       es.state_data "
        "FROM executions e "
        "LEFT JOIN execution_states es ON es.execution_id = e.id "
        "WHERE e.workflow_id = :wid AND e.status = 'completed' "
        "ORDER BY e.completed_at DESC LIMIT 1"
    ), {"wid": workflow_id})).fetchone()

    if not row or not row[5]:
        raise HTTPException(status_code=404, detail="No results available yet. Run the workflow first.")

    exec_id = str(row[0])
    started_at = row[2]
    completed_at = row[3]
    state_data = row[5] if isinstance(row[5], dict) else _json.loads(row[5])

    # ── Extract only the richest data set (last step with a data array) ──
    all_items = []
    steps_info = state_data.get("steps", {})
    sorted_steps = sorted(steps_info.keys())  # step-1, step-2, ...
    for step_id in reversed(sorted_steps):
        step_info = steps_info[step_id]
        if not isinstance(step_info, dict):
            continue
        output = step_info.get("output")
        if isinstance(output, dict) and "data" in output and isinstance(output["data"], list) and output["data"]:
            all_items = output["data"]
            break
        elif isinstance(output, list) and output and isinstance(output[0], dict):
            all_items = output
            break

    # ── Define clean column order and human-readable headers ──
    COLUMN_CONFIG = [
        ("ean",             "EAN/GTIN",         16),
        ("title",           "Model",            45),
        ("quantity",        "Quantity",          12),
        ("asin",            "ASIN",             14),
        ("deal_price",      "Amazon Price (€)", 18),
        ("amazon_price_chf","Amazon (CHF)",     14),
        ("galaxus_price",   "Galaxus (CHF)",    14),
        ("price_diff",      "Difference (%)",   15),
        ("cheaper_at",      "Cheaper At",       14),
        ("galaxus_found",   "Galaxus Match",    14),
        ("rating",          "Rating",           10),
        ("bsr",             "BSR",              10),
        ("amazon_url",      "Amazon URL",       38),
        ("galaxus_url",     "Galaxus URL",      38),
        ("galaxus_title",   "Galaxus Title",    35),
        ("search_query",    "Search Query",     25),
        ("image",           "Image URL",        40),
    ]

    # Filter columns: keep if exists in data OR is a placeholder column (quantity)
    ALWAYS_SHOW = {"ean", "title", "quantity", "asin"}  # Core columns always present
    if all_items:
        existing_keys = set()
        for item in all_items:
            if isinstance(item, dict):
                existing_keys.update(item.keys())
        columns = [(k, h, w) for k, h, w in COLUMN_CONFIG if k in existing_keys or k in ALWAYS_SHOW]
        # Add any extra keys not in config at the end
        configured_keys = {k for k, _, _ in COLUMN_CONFIG}
        for key in sorted(existing_keys - configured_keys):
            columns.append((key, key.replace("_", " ").title(), 20))
    else:
        columns = []

    # ── Build Excel workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    data_font = Font(name="Arial", size=10)
    number_fmt_pct = '0.0"%"'
    number_fmt_chf = '#,##0.00'
    green_font = Font(name="Arial", size=10, color="1B7A2B")
    red_font = Font(name="Arial", size=10, color="CC0000")
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    bool_true_fill = PatternFill("solid", fgColor="E8F5E9")
    bool_false_fill = PatternFill("solid", fgColor="FFEBEE")

    if columns and all_items:
        # Write header row
        for col_idx, (key, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = width

        # Write data rows
        for row_idx, item in enumerate(all_items, 2):
            if not isinstance(item, dict):
                continue
            for col_idx, (key, header, _w) in enumerate(columns, 1):
                val = item.get(key, "")
                if isinstance(val, (dict, list)):
                    val = _json.dumps(val, ensure_ascii=False, default=str)

                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                # Alternating row fill
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

                # ── Column-specific formatting ──
                if key == "price_diff" and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = '0.0"%"'
                    cell.font = green_font if val >= 0 else red_font
                elif key in ("amazon_price_chf", "galaxus_price") and isinstance(val, (int, float)):
                    cell.value = val
                    cell.number_format = '#,##0.00'
                    cell.font = data_font
                elif key == "galaxus_found":
                    cell.value = "Yes" if val else "No"
                    cell.font = data_font
                    if val:
                        cell.fill = bool_true_fill
                    else:
                        cell.fill = bool_false_fill
                elif key in ("amazon_url", "galaxus_url", "image") and val:
                    cell.value = val
                    cell.font = link_font
                    cell.hyperlink = str(val)
                elif key == "cheaper_at":
                    cell.value = val
                    cell.font = green_font if val == "Amazon" else red_font if val == "Galaxus" else data_font
                else:
                    cell.value = val
                    cell.font = data_font

        # Freeze header row + auto-filter
        ws.freeze_panes = "A2"
        last_col = ws.cell(row=1, column=len(columns)).column_letter
        ws.auto_filter.ref = f"A1:{last_col}{len(all_items) + 1}"

    # ── Info sheet ──
    info_ws = wb.create_sheet("Info")
    meta_bold = Font(name="Arial", size=11, bold=True)
    meta_font = Font(name="Arial", size=11)
    meta = [
        ("Workflow", wf.name),
        ("Execution ID", exec_id),
        ("Started", started_at.strftime("%d/%m/%Y %H:%M") if started_at else "—"),
        ("Completed", completed_at.strftime("%d/%m/%Y %H:%M") if completed_at else "—"),
        ("Total Records", len(all_items)),
        ("Exported At", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]
    for r, (label, value) in enumerate(meta, 1):
        info_ws.cell(row=r, column=1, value=label).font = meta_bold
        info_ws.cell(row=r, column=2, value=value).font = meta_font
    info_ws.column_dimensions["A"].width = 18
    info_ws.column_dimensions["B"].width = 55

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)

    safe_name = wf.name.replace(" ", "_")
    return FileResponse(
        path=tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{safe_name}_results.xlsx",
    )


@router.get("/workflows/{workflow_id}/detail")
async def get_workflow_detail(
    workflow_id: str,
    current_user: TokenPayload = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get full workflow detail for the dashboard page:
    workflow info, latest execution, results summary, schedules.
    """
    from sqlalchemy import text as sa_text

    svc = WorkflowService(db)
    wf = await svc.get_by_id_and_org(workflow_id, current_user.org_id)
    if not wf:
        raise HTTPException(status_code=404, detail="Workflow not found")

    # Latest execution
    row = (await db.execute(sa_text(
        "SELECT id, status, trigger_type, started_at, completed_at, duration_ms, error_message "
        "FROM executions WHERE workflow_id = :wid ORDER BY created_at DESC LIMIT 1"
    ), {"wid": workflow_id})).fetchone()

    latest_execution = None
    if row:
        latest_execution = {
            "id": str(row[0]),
            "status": row[1],
            "trigger_type": row[2],
            "started_at": row[3].isoformat() if row[3] else None,
            "completed_at": row[4].isoformat() if row[4] else None,
            "duration_ms": row[5],
            "error_message": row[6],
        }

    # Total execution count
    count_row = (await db.execute(sa_text(
        "SELECT COUNT(*) FROM executions WHERE workflow_id = :wid"
    ), {"wid": workflow_id})).fetchone()
    total_executions = count_row[0] if count_row else 0

    # Schedules for this workflow
    schedule_rows = (await db.execute(sa_text(
        "SELECT id, name, cron_expression, timezone, is_enabled, next_run_at "
        "FROM schedules WHERE workflow_id = :wid AND is_deleted = false ORDER BY created_at"
    ), {"wid": workflow_id})).fetchall()

    schedules = [{
        "id": str(r[0]),
        "name": r[1],
        "cron_expression": r[2],
        "timezone": r[3],
        "is_enabled": r[4],
        "next_run_at": r[5].isoformat() if r[5] else None,
    } for r in schedule_rows]

    # Latest results summary — from DB execution_states (authoritative)
    # Use latest COMPLETED execution (not just latest — a running execution has no state yet)
    results_summary = None
    completed_row = (await db.execute(sa_text(
        "SELECT id, status, trigger_type, started_at, completed_at, duration_ms "
        "FROM executions WHERE workflow_id = :wid AND status = 'completed' "
        "ORDER BY completed_at DESC LIMIT 1"
    ), {"wid": workflow_id})).fetchone()

    if completed_row:
        state_row = (await db.execute(sa_text(
            "SELECT state_data FROM execution_states WHERE execution_id = :eid LIMIT 1"
        ), {"eid": str(completed_row[0])})).fetchone()

        if state_row and state_row[0]:
            import json as _json
            sd = state_row[0] if isinstance(state_row[0], dict) else _json.loads(state_row[0])
            total_items = 0
            # Use the richest step (last step with data array)
            steps = sd.get("steps", {})
            if isinstance(steps, dict):
                for sid in sorted(steps.keys(), reverse=True):
                    si = steps[sid]
                    if not isinstance(si, dict):
                        continue
                    out = si.get("output")
                    if isinstance(out, dict) and "data" in out and isinstance(out["data"], list) and out["data"]:
                        total_items = len(out["data"])
                        break
                    elif isinstance(out, list) and out and isinstance(out[0], dict):
                        total_items = len(out)
                        break

            results_summary = {
                "saved_at": completed_row[4].isoformat() if completed_row[4] else (completed_row[3].isoformat() if completed_row[3] else None),
                "execution_id": str(completed_row[0]),
                "total_items": total_items,
                "file_size": 0,  # Size will be calculated on download
            }

    return {
        "workflow": {
            "id": str(wf.id),
            "name": wf.name,
            "description": wf.description,
            "status": wf.status,
            "version": wf.version,
            "is_enabled": wf.is_enabled,
            "created_at": wf.created_at.isoformat() if wf.created_at else None,
            "updated_at": wf.updated_at.isoformat() if wf.updated_at else None,
            "step_count": len(wf.definition.get("steps", [])) if wf.definition else 0,
        },
        "latest_execution": latest_execution,
        "total_executions": total_executions,
        "schedules": schedules,
        "results_summary": results_summary,
    }


@router.get("/stats")
async def storage_stats(
    current_user: TokenPayload = Depends(get_current_active_user),
):
    """Get overall storage statistics."""
    storage = get_storage_service()
    return storage.get_storage_stats()
